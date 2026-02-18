"""
Shopify Deep Scanner - Subscription App Detector
- Scans homepage, product pages, collections, and all linked pages
- Detects 50+ subscription apps via technical selectors, proxy paths, data attributes, UI strings
- GitHub Actions compatible with chunk-based processing (CHUNK_INDEX / CHUNK_TOTAL env vars)
"""

import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
import time
import random
import os
import re
import json
from bs4 import BeautifulSoup

# ─────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────
INPUT_FILE  = os.getenv("INPUT_FILE", "SKU Subscription Data.csv")
OUTPUT_FILE = os.getenv("OUTPUT_FILE", "Shopify_Deep_Analysis.xlsx")
THREADS     = int(os.getenv("THREADS", "5"))
TIMEOUT     = (6, 12)

# GitHub Actions chunking: set CHUNK_INDEX (0-based) and CHUNK_TOTAL
CHUNK_INDEX = int(os.getenv("CHUNK_INDEX", "0"))
CHUNK_TOTAL = int(os.getenv("CHUNK_TOTAL", "1"))

# ─────────────────────────────────────────
# APP SIGNATURES (from your table + extras)
# ─────────────────────────────────────────
APP_SIGNATURES = [
    # (App Name, [selectors/keywords to look for in HTML])
    ("Recharge Subscriptions",      ["rc_container", "/apps/recharge/", "data-recharge-provider", "Subscribe & Save", "recharge"]),
    ("Bold Subscriptions",          ["bold-ro__product", "/apps/subscriptions/", "bold_recurring_id", "Recurring Order", "boldapps"]),
    ("Appstle Subscriptions",       ["appstle_init", "/apps/appstle-subscriptions/", "data-appstle-plan", "Subscription Management", "appstle"]),
    ("Seal Subscriptions",          ["seal-subs", "/apps/seal-subscriptions/", "data-seal-id", "Auto-ship", "seal-subscriptions"]),
    ("Skio Subscriptions",          ["skio-plan-picker", "/a/skio/", "data-skio-plan-id", "Passwordless Login", "skio"]),
    ("Loop Subscriptions",          ["loop-subscription-widget", "/a/loop_subscriptions/", "data-loop-id", "Delivered Monthly", "loopwork"]),
    ("Stay AI",                     ["stay-ai-widget", "/a/stay/", "data-stay-plan", "Retention Engine", "stayai"]),
    ("Ordergroove",                 ["og-offer", "/apps/ordergroove/", "data-og-module", "Subscription Link", "ordergroove"]),
    ("Smartrr",                     ["smartrr-widget", "/a/smartrr/", "data-smartrr-id", "Account Portal", "smartrr"]),
    ("PayWhirl",                    ["paywhirl-widget", "/apps/paywhirl/", "data-paywhirl-id", "Billing Portal", "paywhirl"]),
    ("Ongoing Subscriptions",       ["ongoing-subscription-widget", "/apps/ongoing/", "ongoing_id", "Automatic Billing"]),
    ("Subify",                      ["subify-subscription-widget", "/apps/subify/", "data-subify-plan", "Periodic Discount", "subify"]),
    ("Native Shopify Subscriptions",["selling_plan", "selling_plan_id", "selling_plan_groups"]),
    ("Recurpay",                    ["recurpay-widget", "/apps/recurpay/", "data-recurpay-id", "Self Service", "recurpay"]),
    ("Propel Subscriptions",        ["propel-widget", "/apps/propel/", "data-propel-plan", "Fixed Price"]),
    ("Monto Subscriptions",         ["monto-subscription-widget", "/apps/monto/", "data-monto-plan", "Recurring Logic"]),
    ("Simple Subscriptions",        ["simple-sub-widget", "/apps/simple-sub/", "data-simple-plan", "Billing Interval"]),
    ("CASA Subscriptions",          ["casa-widget", "/a/casa/", "data-casa-plan", "Direct-to-consumer"]),
    ("Gronos Subscriptions",        ["gronos-widget", "/apps/gronos/", "data-gronos-id", "Simple Setup"]),
    ("Subbly",                      ["subbly-checkout", "/a/subbly/", "data-subbly-id", "Checkout Builder", "subbly"]),
    ("ChargeBee",                   ["chargebee-widget", "/apps/chargebee/", "data-cb-plan-id", "Enterprise Billing", "chargebee"]),
    ("Recurly",                     ["recurly-widget", "/apps/recurly/", "data-recurly-id", "Revenue Recovery", "recurly"]),
    ("Spur Subscriptions",          ["spur-widget", "/a/spur/", "data-spur-id", "Mobile Optimized"]),
    ("Beboxed",                     ["beboxed-widget", "/apps/beboxed/", "data-beboxed-id", "Curation"]),
    ("Upscribe",                    ["upscribe-widget", "/a/upscribe/", "data-upscribe-id", "LTV Tracking", "upscribe"]),
    ("Zest Subscriptions",          ["zest-widget", "/a/zest/", "data-zest-id", "Food & Beverage"]),
    ("Klaviyo Subscriptions",       ["klaviyo-sub-widget", "/apps/klaviyo/", "data-klaviyo-id", "Email Logic"]),
    ("ChargeZen",                   ["chargezen-widget", "/a/chargezen/", "data-chargezen-id", "Optimization"]),
    ("Retentio",                    ["retentio-widget", "/a/retentio/", "data-retentio-id", "Exit Intent"]),
    ("Subscrimo",                   ["subscrimo-widget", "/a/subscrimo/", "data-subscrimo-id", "Visual Editor"]),
    ("Growave Subscriptions",       ["growave-sub-widget", "/apps/growave/", "data-growave-id", "Recurring Points", "growave"]),
    ("Yotpo Subscriptions",         ["yotpo-sub-widget", "/apps/yotpo/", "data-yotpo-id", "Subscription Rewards", "yotpo"]),
    ("Smile Subscriptions",         ["smile-sub-widget", "/apps/smile/", "data-smile-id", "Referral Plan"]),
    ("Rivo Subscriptions",          ["rivo-sub-widget", "/a/rivo/", "data-rivo-id", "Retention Platform", "rivo"]),
    ("LoyaltyLion Subscriptions",   ["lion-sub-widget", "/apps/loyaltylion/", "data-lion-id", "Reward Strategy"]),
    ("Rebuy Subscriptions",         ["rebuy-sub-widget", "/apps/rebuy/", "data-rebuy-id", "AI Recommendations", "rebuy"]),
    ("Hulk Subscriptions",          ["hulk-sub-widget", "/apps/hulk-apps/", "data-hulk-id", "Recurring Discount"]),
    ("Vitals Subscriptions",        ["vitals-sub-widget", "/apps/vitals/", "data-vitals-id", "Subscription Logic", "vitals"]),
    ("EasySub",                     ["easysub-widget", "/a/easysub/", "data-easysub-id", "One-click Billing"]),
    ("Prime Subscriptions",         ["prime-sub-widget", "/apps/prime/", "data-prime-id", "Digital Products"]),
    ("Subscription Plus",           ["plus-sub-widget", "/apps/subplus/", "data-plus-id", "Trial Management"]),
    ("QPilot",                      ["qpilot-widget", "/apps/qpilot/", "data-qpilot-id", "Autoship Cloud", "qpilot"]),
    ("Subflow",                     ["subflow-widget", "/a/subflow/", "data-subflow-id", "Flow Management"]),
    ("Kaching Subscriptions",       ["kaching-widget", "/a/kaching/", "data-kaching-id", "Billing Logic"]),
    ("Plobal Subscriptions",        ["plobal-sub-widget", "/apps/plobal/", "data-plobal-id", "Mobile App"]),
    ("Tapcart Subscriptions",       ["tapcart-sub-widget", "/apps/tapcart/", "data-tapcart-id", "Mobile Commerce"]),
    # Extra generic keywords
    ("Generic Subscription",        ["subscribe-and-save", "subscription-widget", "subscription_form", "subscription-plan",
                                     "auto-renew", "autorenew", "subscribe_button", "recurring-billing",
                                     "membership-plan", "subscribe-save", "subscribe--save"]),
]

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
]

def get_headers():
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
    }

def fetch_url(url, retries=2):
    """Fetch a URL with retry logic. Returns (html_text, status_code)."""
    for attempt in range(retries + 1):
        try:
            r = requests.get(url, headers=get_headers(), timeout=TIMEOUT, allow_redirects=True)
            if r.status_code == 429:
                time.sleep(3 * (attempt + 1))
                continue
            return r.text, r.status_code
        except Exception as e:
            if attempt == retries:
                return "", str(e)[:40]
            time.sleep(1)
    return "", 429

def detect_apps_in_html(html: str) -> list:
    """Check HTML against all app signatures. Returns list of detected app names."""
    html_lower = html.lower()
    detected = []
    for app_name, keywords in APP_SIGNATURES:
        for kw in keywords:
            if kw.lower() in html_lower:
                detected.append(app_name)
                break
    return detected

def get_internal_links(html: str, base_domain: str, limit: int = 20) -> list:
    """Extract internal page links from HTML (collections, pages, products)."""
    try:
        soup = BeautifulSoup(html, "html.parser")
    except Exception:
        return []
    links = set()
    for tag in soup.find_all("a", href=True):
        href = tag["href"]
        # Keep only internal paths for collections, pages, products
        if href.startswith("/") and not href.startswith("//"):
            if any(p in href for p in ["/collections/", "/pages/", "/products/", "/blogs/"]):
                full = f"https://{base_domain}{href.split('?')[0]}"
                links.add(full)
        elif base_domain in href:
            if any(p in href for p in ["/collections/", "/pages/", "/products/", "/blogs/"]):
                links.add(href.split("?")[0])
    return list(links)[:limit]

def get_all_products_api(domain) -> tuple:
    """Paginate /products.json - returns (products_list, status)."""
    all_products = []
    page = 1
    while True:
        html, status = fetch_url(f"https://{domain}/products.json?limit=250&page={page}")
        if status != 200:
            return all_products, status
        try:
            products = json.loads(html).get("products", [])
        except Exception:
            return all_products, "json_error"
        if not products:
            return all_products, 200
        all_products.extend(products)
        if len(products) < 250:
            return all_products, 200
        page += 1
        time.sleep(random.uniform(0.3, 0.8))

def scan_store(domain: str) -> dict:
    """Full store scan: homepage + pages + product API + product JS pages."""
    domain = domain.strip().lower().replace("https://","").replace("http://","").split("/")[0]
    if not domain:
        return {"domain": domain, "status": "empty"}

    time.sleep(random.uniform(0.1, 0.4))

    result = {
        "Domain": domain,
        "Status": "unknown",
        "Subscription_App": "",
        "Apps_Detected": "",
        "Page_Found_On": "",
        "Total_SKUs": 0,
        "Subscription_Products": 0,
        "Sub_Product_Names": "",
        "Sub_Plan_Names": "",
        "Sub_Product_Links": "",
        "Pages_Scanned": 0,
    }

    all_detected_apps = {}   # {app_name: [pages found on]}
    pages_scanned = 0

    # ── 1. Homepage ──────────────────────────────────────────
    home_html, home_status = fetch_url(f"https://{domain}")
    if not home_html:
        result["Status"] = f"blocked_{home_status}"
        return result

    pages_scanned += 1
    home_apps = detect_apps_in_html(home_html)
    for app in home_apps:
        all_detected_apps.setdefault(app, []).append("homepage")

    # ── 2. Internal links from homepage ──────────────────────
    internal_links = get_internal_links(home_html, domain, limit=25)
    # Prioritise pages and collections first, then products
    priority = [l for l in internal_links if "/pages/" in l or "/collections/" in l]
    rest     = [l for l in internal_links if l not in priority]
    scan_links = (priority + rest)[:20]

    for url in scan_links:
        time.sleep(random.uniform(0.2, 0.6))
        html, status = fetch_url(url)
        if html:
            pages_scanned += 1
            page_apps = detect_apps_in_html(html)
            page_label = url.replace(f"https://{domain}", "")
            for app in page_apps:
                all_detected_apps.setdefault(app, []).append(page_label)

    # ── 3. Shopify product API ────────────────────────────────
    products, api_status = get_all_products_api(domain)
    result["Total_SKUs"] = len(products)

    sub_products = []
    checked = 0
    for p in products:
        if checked >= 30:   # max 30 product .js checks per store
            break
        time.sleep(random.uniform(0.1, 0.3))
        js_html, js_status = fetch_url(f"https://{domain}/products/{p['handle']}.js")
        if js_status == 200:
            try:
                data = json.loads(js_html)
                plans = data.get("selling_plan_groups", [])
                if plans:
                    sub_products.append({
                        "title": data["title"],
                        "price": data.get("price", 0) / 100,
                        "plans": ", ".join([pl["name"] for pl in plans]),
                        "link": f"https://{domain}/products/{p['handle']}"
                    })
                    # Also scan product HTML for app detection
                    prod_html, _ = fetch_url(f"https://{domain}/products/{p['handle']}")
                    if prod_html:
                        pages_scanned += 1
                        prod_apps = detect_apps_in_html(prod_html)
                        for app in prod_apps:
                            all_detected_apps.setdefault(app, []).append(f"/products/{p['handle']}")
            except Exception:
                pass
        checked += 1

    # ── 4. Build result ───────────────────────────────────────
    result["Pages_Scanned"] = pages_scanned
    result["Subscription_Products"] = len(sub_products)

    # Remove "Generic Subscription" if a named app is found
    named_apps = {k: v for k, v in all_detected_apps.items() if k != "Generic Subscription"}
    if not named_apps:
        named_apps = all_detected_apps  # fallback to generic

    if named_apps:
        # Primary app = most pages detected on
        primary = max(named_apps, key=lambda k: len(named_apps[k]))
        result["Subscription_App"]  = primary
        result["Apps_Detected"]     = " | ".join(named_apps.keys())
        result["Page_Found_On"]     = " | ".join(
            f"{app}: {', '.join(set(pages[:3]))}"
            for app, pages in named_apps.items()
        )

    if sub_products:
        result["Sub_Product_Names"] = " | ".join(p["title"] for p in sub_products[:10])
        result["Sub_Plan_Names"]    = " | ".join(p["plans"] for p in sub_products[:10])
        result["Sub_Product_Links"] = " | ".join(p["link"] for p in sub_products[:10])
        result["Status"] = "found"
    elif named_apps:
        result["Status"] = "app_detected_no_product_api"
    else:
        result["Status"] = "no_subscription"

    return result

# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────
def get_url_column(df):
    for col in df.columns:
        if any(w in col.lower() for w in ["url","domain","store","site","link","web","company"]):
            return col
    return df.columns[0]

def clean_domain(d):
    return str(d).strip().lower().replace("https://","").replace("http://","").split("/")[0]

# ─────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────
def main():
    print(f"📥 Loading: {INPUT_FILE}", flush=True)
    df_input = pd.read_csv(INPUT_FILE)
    url_col  = get_url_column(df_input)
    domains  = [clean_domain(d) for d in df_input[url_col].dropna().tolist()]
    domains  = [d for d in domains if d]

    # ── Chunking for GitHub Actions parallel runs ─────────────
    if CHUNK_TOTAL > 1:
        chunk_size = len(domains) // CHUNK_TOTAL
        start = CHUNK_INDEX * chunk_size
        end   = start + chunk_size if CHUNK_INDEX < CHUNK_TOTAL - 1 else len(domains)
        domains = domains[start:end]
        print(f"🔀 Chunk {CHUNK_INDEX+1}/{CHUNK_TOTAL}: domains {start}–{end} ({len(domains)} stores)", flush=True)

    print(f"🚀 Scanning {len(domains)} stores | threads={THREADS}\n", flush=True)

    all_results = []
    completed = 0

    with ThreadPoolExecutor(max_workers=THREADS) as executor:
        futures = {executor.submit(scan_store, d): d for d in domains}
        with tqdm(total=len(domains), dynamic_ncols=True) as pbar:
            for future in as_completed(futures, timeout=72000):
                try:
                    r = future.result(timeout=90)
                    all_results.append(r)
                except Exception as e:
                    domain = futures[future]
                    all_results.append({"Domain": str(domain), "Status": f"timeout_{str(e)[:30]}"})
                completed += 1
                pbar.update(1)
                if completed % 50 == 0:
                    found    = sum(1 for r in all_results if r.get("Status") in ["found","app_detected_no_product_api"])
                    blocked  = sum(1 for r in all_results if "blocked" in r.get("Status",""))
                    tqdm.write(f"[{completed}/{len(domains)}] ✅ With sub data: {found} | ❌ Blocked: {blocked}")

    # ── Summary ───────────────────────────────────────────────
    df = pd.DataFrame(all_results)
    print("\n─── STATUS SUMMARY ───", flush=True)
    print(df["Status"].value_counts().to_string(), flush=True)
    print(f"\n📊 Total subscription stores: {df[df['Status'].isin(['found','app_detected_no_product_api'])].shape[0]}", flush=True)

    # ── App frequency ─────────────────────────────────────────
    app_counts = {}
    for apps_str in df["Apps_Detected"].dropna():
        for app in apps_str.split(" | "):
            if app:
                app_counts[app] = app_counts.get(app, 0) + 1
    df_apps = pd.DataFrame(sorted(app_counts.items(), key=lambda x: -x[1]), columns=["App","Store_Count"])

    # ── Write Excel ───────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    def write_sheet(wb, df_data, sheet_name, header_color="1F4E79"):
        ws = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(df_data, index=False, header=True):
            ws.append(r)
        header_fill = PatternFill("solid", start_color=header_color)
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=9)
        return ws

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Sheet 1: All Results
    write_sheet(wb, df, "All_Stores", "1F4E79")

    # Sheet 2: Only stores with subscriptions
    df_found = df[df["Status"].isin(["found","app_detected_no_product_api"])].copy()
    if not df_found.empty:
        write_sheet(wb, df_found, "Subscription_Stores", "375623")

    # Sheet 3: App frequency
    write_sheet(wb, df_apps, "App_Usage_Stats", "7030A0")

    # Sheet 4: Status log compact
    df_status = df[["Domain","Status","Subscription_App","Total_SKUs","Subscription_Products","Pages_Scanned"]].copy()
    write_sheet(wb, df_status, "Status_Log", "843C0C")

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved: {OUTPUT_FILE}", flush=True)

if __name__ == "__main__":
    main()
