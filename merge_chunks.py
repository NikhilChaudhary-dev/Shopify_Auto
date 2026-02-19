"""
merge_chunks.py - Jo bhi chunks mile unhe merge karo (missing chunks skip)
"""

import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

OUTPUT_FILE = "FINAL_Shopify_Deep_Analysis.xlsx"

def write_sheet(wb, df_data, sheet_name, header_color="1F4E79"):
    if df_data.empty:
        return
    ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df_data, index=False, header=True):
        ws.append(r)
    header_fill = PatternFill("solid", start_color=header_color)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=9)

def main():
    # Har jagah se files dhundo
    search_paths = [
        "./chunks/output_chunk_*.xlsx",
        "./output_chunk_*.xlsx",
        "./**/output_chunk_*.xlsx",
        "./chunks/**/output_chunk_*.xlsx",
    ]

    files = []
    for pattern in search_paths:
        found = glob.glob(pattern, recursive=True)
        files.extend(found)

    # Duplicates remove karo
    files = list(set(files))

    if not files:
        # Last resort — saari xlsx files dhundo
        files = glob.glob("./**/*.xlsx", recursive=True)
        files = [f for f in files if "chunk" in f.lower()]

    files = sorted(files)
    print(f"📦 Found {len(files)} chunk files:", flush=True)
    for f in files:
        print(f"   {f}", flush=True)

    if not files:
        print("❌ Koi chunk file nahi mili!", flush=True)
        # Empty file banao taaki artifact upload fail na ho
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "No chunk files found"
        wb.save(OUTPUT_FILE)
        return

    # ── Saare chunks read karo ────────────────────────────
    all_dfs        = []
    all_status_dfs = []
    missing_chunks = []

    for i in range(20):  # 0 to 19
        chunk_file = next((f for f in files if f"chunk_{i}" in f or f"chunk-{i}" in f), None)
        if not chunk_file:
            missing_chunks.append(i)
            print(f"  ⚠️  chunk-{i}: MISSING", flush=True)
            continue
        try:
            df = pd.read_excel(chunk_file, sheet_name="Subscription_Products")
            all_dfs.append(df)
            print(f"  ✅ chunk-{i}: {len(df)} subscription products", flush=True)
        except Exception:
            # Sheet nahi hai matlab koi subscription nahi mili us chunk mein
            print(f"  ℹ️  chunk-{i}: No subscription products sheet", flush=True)

        try:
            df_log = pd.read_excel(chunk_file, sheet_name="Status_Log")
            all_status_dfs.append(df_log)
        except Exception:
            pass

    if missing_chunks:
        print(f"\n⚠️  Missing chunks: {missing_chunks}", flush=True)
        print(f"   In stores ka data nahi aaya. Dobara run karo agar chahiye.\n", flush=True)

    # ── Merge karo ───────────────────────────────────────
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if all_dfs:
        df_all = pd.concat(all_dfs, ignore_index=True)
        print(f"\n📊 Total subscription products: {len(df_all)}", flush=True)

        # Store summary
        summary = []
        for store, grp in df_all.groupby("Store"):
            summary.append({
                "Store":                 store,
                "Total_SKUs":            grp["Total_SKUs"].iloc[0],
                "Subscription_Products": len(grp),
                "Ratio":                 f"{len(grp)}/{grp['Total_SKUs'].iloc[0]}",
                "Plan_Names":            " | ".join(grp["Sub_Plans"].unique()[:5]),
                "Product_Names":         " | ".join(grp["Product_Title"].tolist()[:10])
            })
        df_summary = pd.DataFrame(summary)
        print(f"🏪 Total subscription stores: {len(df_summary)}", flush=True)

        write_sheet(wb, df_all,     "All_Subscription_Products", "1F4E79")
        write_sheet(wb, df_summary, "Store_Summary",              "375623")
    else:
        print("\n⚠️  Kisi bhi chunk mein subscription products nahi mili", flush=True)
        ws = wb.create_sheet("No_Data")
        ws['A1'] = "No subscription products found in any chunk"

    # Status log merge
    if all_status_dfs:
        df_log_all = pd.concat(all_status_dfs, ignore_index=True)
        write_sheet(wb, df_log_all, "Status_Log", "843C0C")

        # Status summary
        df_status_summary = df_log_all["Status"].value_counts().reset_index()
        df_status_summary.columns = ["Status", "Count"]
        write_sheet(wb, df_status_summary, "Status_Summary", "595959")

        print(f"\n─── FINAL STATUS SUMMARY ───", flush=True)
        print(df_status_summary.to_string(index=False), flush=True)

    # Missing chunks info
    if missing_chunks:
        ws_missing = wb.create_sheet("Missing_Chunks")
        ws_missing['A1'] = "Missing Chunk Numbers"
        ws_missing['B1'] = "Stores Range"
        for row, chunk_num in enumerate(missing_chunks, start=2):
            start = chunk_num * (6600 // 20)
            end   = start + (6600 // 20)
            ws_missing.cell(row=row, column=1, value=f"chunk-{chunk_num}")
            ws_missing.cell(row=row, column=2, value=f"{start}–{end}")

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved: {OUTPUT_FILE}", flush=True)

if __name__ == "__main__":
    main()
